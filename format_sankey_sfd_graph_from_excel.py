import openpyxl
import sys
import pprint
import csv

def get_data(path, sheet_index, min_row):
    wb = openpyxl.load_workbook(path)
    available_sheets = wb.sheetnames
    thename = available_sheets[sheet_index]
    sheet = wb[thename]

    result = []
    for row in sheet.iter_rows(min_row=min_row):
        r = [c.value for c in row]
        result.append(
            {'feature': r[3],
             'feature_score': r[1],
             'feature_count': r[0]}
        )
    return result

# foo1 = get_data("sample-excel/pairing1/Military-c SPD.xlsx", 1, 2)
# foo2 = get_data("sample-excel/pairing1/MaritimeSPD.xlsx", 0, 2)

foo2 = get_data("sample-excel/pairing2/Chase SPD.xlsx", 1, 3)
foo1 = get_data("sample-excel/pairing2/RiverChasePx4 SPD.xlsx", 1, 3)

features = {}

def add_to_features(records):
    for rec in records:
        features.setdefault(rec['feature'], []).append(rec)


add_to_features(foo1)
add_to_features(foo2)


# writer = csv.writer(sys.stdout)
# writer.writerow(['feature', 'score1', 'count1', 'score2', 'count2', 'scoredelta', 'countdelta'])

all_nodes = []
all_links = []

for feature, records in features.items():
    if len(records) > 1:
        rec1 = records[0]
        rec2 = records[1]

        score1 = float(rec1['feature_score'])
        score2 = float(rec2['feature_score'])
        count1 = float(rec1['feature_count'])
        count2 = float(rec2['feature_count'])

        id1 = len(all_nodes)
        id2 = id1 + 1

        all_nodes.append(
            {'name': feature,
             'score': score1,
             'count': count1}
        )
        all_nodes.append(
            {'name': feature,
             'score': score2,
             'count': count2}
        )

        countdelta = count2 - count1

        all_links.append(
            {'source': id1,
             'target': id2,
             'value': countdelta}
        )

        # csv_record = [
        #     feature,
        #     score1,
        #     count1,
        #     score2,
        #     count2,
        #     score2 - score1,
        #     count2 - count1
        # ]

        # writer.writerow(csv_record)

pprint.pprint(
    {'nodes': all_nodes, 'links': all_links}
)
